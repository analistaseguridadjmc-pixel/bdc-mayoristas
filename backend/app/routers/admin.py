from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional
from datetime import date
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import get_db
from app.middleware.security import require_role, hashear_password

router = APIRouter(prefix="/admin", tags=["Admin"])

ROLES_ADMIN = ["jdt", "sdt", "jvo", "auditor"]

# ── DASHBOARD ─────────────────────────────────────────────────
@router.get("/dashboard")
async def dashboard(tienda_codigo: Optional[str] = Query(None),
                    db: AsyncSession = Depends(get_db),
                    usuario=Depends(require_role(ROLES_ADMIN))):
    tc = tienda_codigo or None
    p = {"tc": tc}

    r1 = await db.execute(text("""
        SELECT
            COUNT(*) FILTER (WHERE vm.estado NOT IN ('entregada','vencida')) AS activas,
            COUNT(*) FILTER (WHERE vm.estado = 'vencida') AS vencidas,
            COUNT(*) FILTER (WHERE vm.estado = 'entregada') AS entregadas
        FROM ventas_mayoristas vm
        JOIN tiendas t ON t.id = vm.tienda_id
        WHERE (:tc IS NULL OR t.codigo = :tc)
    """), p)
    totales = dict(r1.mappings().first())

    r2 = await db.execute(text("""
        SELECT COUNT(*) FROM retiros ret
        JOIN formatos_gvm001 f ON f.id = ret.formato_id
        JOIN ventas_mayoristas vm ON vm.id = f.venta_id
        JOIN tiendas t ON t.id = vm.tienda_id
        WHERE DATE(ret.fecha_retiro AT TIME ZONE 'America/Bogota') = CURRENT_DATE
          AND (:tc IS NULL OR t.codigo = :tc)
    """), p)
    retiros_hoy = r2.scalar() or 0

    r3 = await db.execute(text("""
        SELECT COUNT(*) FROM eventos_auditoria ea
        JOIN usuarios u ON u.id = ea.usuario_id
        JOIN tiendas t ON t.id = u.tienda_id
        WHERE ea.evento LIKE 'ANOMALIA%'
          AND DATE(ea.created_at AT TIME ZONE 'America/Bogota') = CURRENT_DATE
          AND (:tc IS NULL OR t.codigo = :tc)
    """), p)
    anomalias_hoy = r3.scalar() or 0

    r4 = await db.execute(text("""
        SELECT t.nombre AS tienda, t.codigo,
            COUNT(vm.id) FILTER (WHERE vm.estado NOT IN ('entregada','vencida')) AS activas,
            COUNT(vm.id) FILTER (WHERE vm.estado = 'vencida') AS vencidas,
            COUNT(vm.id) FILTER (WHERE vm.estado = 'entregada') AS entregadas
        FROM tiendas t
        LEFT JOIN ventas_mayoristas vm ON vm.tienda_id = t.id
        WHERE t.activa = TRUE AND (:tc IS NULL OR t.codigo = :tc)
        GROUP BY t.id, t.nombre, t.codigo ORDER BY t.codigo
    """), p)
    por_tienda = [dict(r) for r in r4.mappings().all()]

    r5 = await db.execute(text("""
        SELECT vm.numero_factura,
               ROUND(EXTRACT(EPOCH FROM (f.fecha_limite - NOW()))/3600, 1) AS horas_restantes,
               t.nombre AS tienda, c.razon_social AS cliente,
               (3 - COUNT(ret.id)) AS retiros_disponibles
        FROM ventas_mayoristas vm
        JOIN formatos_gvm001 f ON f.venta_id = vm.id AND f.activo = TRUE
        JOIN tiendas t ON t.id = vm.tienda_id
        JOIN clientes_mayoristas c ON c.id = vm.cliente_id
        LEFT JOIN retiros ret ON ret.formato_id = f.id
        WHERE vm.estado NOT IN ('entregada','vencida') AND f.fecha_limite > NOW()
          AND EXTRACT(EPOCH FROM (f.fecha_limite - NOW()))/3600 < 12
          AND (:tc IS NULL OR t.codigo = :tc)
        GROUP BY vm.id, vm.numero_factura, f.fecha_limite, t.nombre, c.razon_social
        ORDER BY horas_restantes ASC LIMIT 10
    """), p)
    proximas_vencer = [dict(r) for r in r5.mappings().all()]

    r6 = await db.execute(text("""
        SELECT ea.evento, ea.created_at, u.nombre AS vigilante,
               t.nombre AS tienda, ea.detalle
        FROM eventos_auditoria ea
        JOIN usuarios u ON u.id = ea.usuario_id
        JOIN tiendas t ON t.id = u.tienda_id
        WHERE ea.evento LIKE 'ANOMALIA%'
          AND (:tc IS NULL OR t.codigo = :tc)
        ORDER BY ea.created_at DESC LIMIT 15
    """), p)
    anomalias = []
    for row in r6.mappings().all():
        d = dict(row)
        d["created_at"] = d["created_at"].isoformat()
        anomalias.append(d)

    return {
        "kpis": {
            "ventas_activas": int(totales.get("activas") or 0),
            "ventas_vencidas": int(totales.get("vencidas") or 0),
            "ventas_entregadas": int(totales.get("entregadas") or 0),
            "retiros_hoy": int(retiros_hoy),
            "anomalias_hoy": int(anomalias_hoy)
        },
        "por_tienda": por_tienda,
        "proximas_vencer": proximas_vencer,
        "anomalias_recientes": anomalias
    }

# ── USUARIOS ──────────────────────────────────────────────────
@router.get("/usuarios")
async def listar_usuarios(db: AsyncSession = Depends(get_db),
                          usuario=Depends(require_role(ROLES_ADMIN))):
    r = await db.execute(text("""
        SELECT u.id, u.nombre, u.documento, u.rol, u.activo, u.created_at,
               t.codigo AS tienda_codigo, t.nombre AS tienda_nombre,
               v.numero_carnet, v.empresa_seguridad, v.turno
        FROM usuarios u
        JOIN tiendas t ON t.id = u.tienda_id
        LEFT JOIN vigilantes v ON v.usuario_id = u.id
        ORDER BY u.rol, t.codigo, u.nombre
    """))
    usuarios = []
    for row in r.mappings().all():
        d = dict(row)
        d["id"] = str(d["id"])
        d["created_at"] = d["created_at"].isoformat()
        usuarios.append(d)
    return usuarios

class CrearUsuarioSchema(BaseModel):
    nombre: str
    documento: str
    password: str
    rol: str
    tienda_codigo: str
    empresa_seguridad: Optional[str] = None
    numero_carnet: Optional[str] = None
    turno: Optional[str] = "diurno"

@router.post("/usuarios")
async def crear_usuario(payload: CrearUsuarioSchema,
                        db: AsyncSession = Depends(get_db),
                        usuario=Depends(require_role(ROLES_ADMIN))):
    roles_validos = ["vigilante", "jdt", "sdt", "jvo", "auditor"]
    if payload.rol not in roles_validos:
        raise HTTPException(status_code=422, detail=f"Rol inválido. Opciones: {roles_validos}")

    dup = await db.execute(text("SELECT id FROM usuarios WHERE documento=:doc"), {"doc": payload.documento})
    if dup.first():
        raise HTTPException(status_code=409, detail={"mensaje": f"Ya existe un usuario con documento {payload.documento}."})

    tienda = await db.execute(text("SELECT id FROM tiendas WHERE codigo=:cod"), {"cod": payload.tienda_codigo})
    t = tienda.first()
    if not t:
        raise HTTPException(status_code=404, detail="Tienda no encontrada.")
    tienda_id = t[0]

    r = await db.execute(text("""
        INSERT INTO usuarios(nombre, documento, password_hash, rol, tienda_id)
        VALUES(:nom, :doc, crypt(:pwd, gen_salt('bf',12)), :rol, :tid)
        RETURNING id
    """), {"nom": payload.nombre, "doc": payload.documento,
           "pwd": payload.password, "rol": payload.rol, "tid": str(tienda_id)})
    nuevo_id = r.scalar()

    if payload.rol == "vigilante":
        if not payload.empresa_seguridad or not payload.numero_carnet:
            await db.rollback()
            raise HTTPException(status_code=422, detail="Vigilantes requieren empresa_seguridad y numero_carnet.")
        await db.execute(text("""
            INSERT INTO vigilantes(usuario_id, empresa_seguridad, numero_carnet, turno)
            VALUES(:uid, :emp, :car, :tur)
        """), {"uid": str(nuevo_id), "emp": payload.empresa_seguridad,
               "car": payload.numero_carnet, "tur": payload.turno or "diurno"})

    await db.commit()
    return {"ok": True, "id": str(nuevo_id), "mensaje": f"Usuario {payload.nombre} creado correctamente."}

@router.patch("/usuarios/{usuario_id}/estado")
async def toggle_usuario(usuario_id: str,
                         db: AsyncSession = Depends(get_db),
                         usuario=Depends(require_role(ROLES_ADMIN))):
    r = await db.execute(text("SELECT activo FROM usuarios WHERE id=:id"), {"id": usuario_id})
    row = r.first()
    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
    nuevo = not row[0]
    await db.execute(text("UPDATE usuarios SET activo=:act WHERE id=:id"), {"act": nuevo, "id": usuario_id})
    if not nuevo:
        await db.execute(text("UPDATE vigilantes SET activo=:act WHERE usuario_id=:id"), {"act": nuevo, "id": usuario_id})
    await db.commit()
    return {"ok": True, "activo": nuevo}

# ── TIENDAS ───────────────────────────────────────────────────
@router.get("/tiendas")
async def listar_tiendas(db: AsyncSession = Depends(get_db),
                         usuario=Depends(require_role(ROLES_ADMIN))):
    r = await db.execute(text("SELECT id, codigo, nombre, ciudad, tipo_caja_fuerte, activa FROM tiendas ORDER BY codigo"))
    return [dict(row) for row in r.mappings().all()]

# ── EXPORTAR EXCEL ────────────────────────────────────────────
def _estilo_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="92610F")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(bottom=Side(style="thin", color="633806"))
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal="center")
        c.border = border

@router.get("/exportar/ventas")
async def exportar_ventas(mes: str = Query(..., description="Formato YYYY-MM"),
                          db: AsyncSession = Depends(get_db),
                          usuario=Depends(require_role(ROLES_ADMIN))):
    try:
        year, month = int(mes.split("-")[0]), int(mes.split("-")[1])
    except:
        raise HTTPException(status_code=422, detail="Formato de mes inválido. Use YYYY-MM")

    r = await db.execute(text("""
        SELECT vm.numero_factura, vm.estado, vm.monto_total,
               vm.requiere_sagrilaft, vm.sagrilaft_diligenciado,
               vm.fecha_venta, vm.jdt_nombre, vm.observaciones,
               t.codigo AS tienda_codigo, t.nombre AS tienda_nombre,
               c.razon_social AS cliente, c.nit,
               u.nombre AS registrado_por,
               v.numero_carnet AS carnet, v.empresa_seguridad AS empresa,
               (SELECT COUNT(*) FROM retiros r2
                JOIN formatos_gvm001 f2 ON f2.id=r2.formato_id
                WHERE f2.venta_id=vm.id) AS total_retiros
        FROM ventas_mayoristas vm
        JOIN tiendas t ON t.id=vm.tienda_id
        JOIN clientes_mayoristas c ON c.id=vm.cliente_id
        LEFT JOIN usuarios u ON u.id=vm.registrado_por_vigilante_id
        LEFT JOIN vigilantes v ON v.usuario_id=vm.registrado_por_vigilante_id
        WHERE EXTRACT(YEAR FROM vm.fecha_venta AT TIME ZONE 'America/Bogota')=:yr
          AND EXTRACT(MONTH FROM vm.fecha_venta AT TIME ZONE 'America/Bogota')=:mo
        ORDER BY vm.fecha_venta DESC
    """), {"yr": year, "mo": month})
    ventas = r.mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas {mes}"
    ws.sheet_view.showGridLines = False

    cols = ["Factura","Estado","Monto Total","Tienda","Cliente","NIT",
            "JDT/SDT","SAGRILAFT","Registrado Por","Carnet","Empresa",
            "Retiros","Fecha","Observaciones"]
    _estilo_header(ws, 1, cols)

    fill_par = PatternFill("solid", fgColor="FDF3E3")
    for i, v in enumerate(ventas, 2):
        row_data = [
            v["numero_factura"], v["estado"], float(v["monto_total"]),
            f"{v['tienda_codigo']} · {v['tienda_nombre']}",
            v["cliente"], v["nit"], v["jdt_nombre"],
            "Sí" if v["sagrilaft_diligenciado"] else ("Requerido" if v["requiere_sagrilaft"] else "No"),
            v["registrado_por"] or "-", v["carnet"] or "-", v["empresa"] or "-",
            int(v["total_retiros"]),
            v["fecha_venta"].strftime("%d/%m/%Y %H:%M"),
            v["observaciones"] or ""
        ]
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                c.fill = fill_par
            if j == 3:
                c.number_format = '$#,##0'

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")), max((len(str(c.value or "")) for c in col), default=0)
        ) + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=ventas_{mes}.xlsx"})

@router.get("/exportar/retiros")
async def exportar_retiros(mes: str = Query(..., description="Formato YYYY-MM"),
                           db: AsyncSession = Depends(get_db),
                           usuario=Depends(require_role(ROLES_ADMIN))):
    try:
        year, month = int(mes.split("-")[0]), int(mes.split("-")[1])
    except:
        raise HTTPException(status_code=422, detail="Formato de mes inválido. Use YYYY-MM")

    r = await db.execute(text("""
        SELECT ret.numero_retiro, ret.tipo, ret.fecha_retiro,
               ret.retira_el_comprador, ret.nombre_autorizado, ret.doc_autorizado,
               ret.carta_autorizacion, ret.sello_colocado,
               ret.vigilante_nombre, ret.vigilante_carnet, ret.vigilante_empresa,
               ret.observaciones,
               vm.numero_factura, vm.monto_total, vm.estado AS estado_venta,
               t.codigo AS tienda_codigo, t.nombre AS tienda_nombre,
               c.razon_social AS cliente, c.nit,
               f.numero_formato
        FROM retiros ret
        JOIN formatos_gvm001 f ON f.id=ret.formato_id
        JOIN ventas_mayoristas vm ON vm.id=f.venta_id
        JOIN tiendas t ON t.id=vm.tienda_id
        JOIN clientes_mayoristas c ON c.id=vm.cliente_id
        WHERE EXTRACT(YEAR FROM ret.fecha_retiro AT TIME ZONE 'America/Bogota')=:yr
          AND EXTRACT(MONTH FROM ret.fecha_retiro AT TIME ZONE 'America/Bogota')=:mo
        ORDER BY ret.fecha_retiro DESC
    """), {"yr": year, "mo": month})
    retiros = r.mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Retiros {mes}"
    ws.sheet_view.showGridLines = False

    cols = ["Factura","Tienda","Cliente","NIT","Formato GVM","Retiro #",
            "Tipo","Fecha Retiro","Retira Comprador","Nombre Autorizado",
            "Doc. Autorizado","Carta Autorización","Sello Colocado",
            "Vigilante","Carnet","Empresa","Observaciones"]
    _estilo_header(ws, 1, cols)

    fill_par = PatternFill("solid", fgColor="FDF3E3")
    for i, r in enumerate(retiros, 2):
        row_data = [
            r["numero_factura"],
            f"{r['tienda_codigo']} · {r['tienda_nombre']}",
            r["cliente"], r["nit"],
            f"GVM-001 #{r['numero_formato']}",
            int(r["numero_retiro"]), r["tipo"],
            r["fecha_retiro"].strftime("%d/%m/%Y %H:%M"),
            "Sí" if r["retira_el_comprador"] else "No",
            r["nombre_autorizado"] or "-", r["doc_autorizado"] or "-",
            "Sí" if r["carta_autorizacion"] else "No",
            "Sí" if r["sello_colocado"] else "No",
            r["vigilante_nombre"] or "-", r["vigilante_carnet"] or "-",
            r["vigilante_empresa"] or "-", r["observaciones"] or ""
        ]
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                c.fill = fill_par

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")), max((len(str(c.value or "")) for c in col), default=0)
        ) + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=retiros_{mes}.xlsx"})
